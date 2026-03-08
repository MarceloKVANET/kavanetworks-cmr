from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os

def crear_excel_cotizacion(datos, nombre_archivo="cotizacion_kvanetworks.xlsx", margen_override=None):
    """
    Toma los datos extraídos por la IA y crea un archivo Excel profesional.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # --- CONFIGURACIÓN DE ESTILOS ---
    azul_kvanetworks = "003366"
    gris_claro = "F2F2F2"
    
    fuente_titulo = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    fuente_negrita = Font(bold=True)
    relleno_azul = PatternFill(start_color=azul_kvanetworks, end_color=azul_kvanetworks, fill_type="solid")
    relleno_gris = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
    
    borde_fino = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # --- ENCABEZADO ---
    ws.merge_cells('A1:E2')
    ws['A1'] = "KVANetworks - COTIZACIÓN TÉCNICA"
    ws['A1'].font = fuente_titulo
    ws['A1'].fill = relleno_azul
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Información básica (Usando getattr por seguridad)
    ws['A4'] = "Fecha:"
    ws['B4'] = datetime.now().strftime("%d/%m/%Y")
    ws['A5'] = "Proyecto:"
    ws['B5'] = getattr(datos, 'resumen_proyecto', "Levantamiento Técnico General")
    
    # --- TABLA DE MATERIALES ---
    ws['A7'] = "Ítem"
    ws['B7'] = "Descripción"
    ws['C7'] = "Cantidad"
    ws['D7'] = "Unidad"
    ws['E7'] = "Costo Unitario ($)" # Cambiado a Costo para mayor claridad
    ws['F7'] = "Margen (%)"
    ws['G7'] = "Total Neto ($)"

    # Aplicar estilos al encabezado de la tabla
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        celda = ws[f'{col}7']
        celda.font = fuente_negrita
        celda.fill = relleno_gris
        celda.border = borde_fino
        celda.alignment = Alignment(horizontal="center")

    # Margen: Usar el manual si existe, si no, usar 30% por defecto (0.30)
    margen_base = margen_override if margen_override is not None else getattr(datos, 'sugerencia_margen', 0.30)

    # Llenar la tabla
    fila_actual = 8
    
    # Si la IA no entregó materiales, creamos una lista vacía para no romper el código
    lista_materiales = getattr(datos, 'materiales_y_servicios', [])
    
    for i, item in enumerate(lista_materiales, 1):
        # Escudos protectores para datos que la IA podría no enviar
        nombre = getattr(item, 'nombre_material', 'Material no especificado')
        desc = getattr(item, 'descripcion_detallada', '')
        nombre_completo = f"{nombre} - {desc}" if desc else nombre
        
        precio_estimado = getattr(item, 'precio_unitario_neto_estimado', 0)
        cantidad = getattr(item, 'cantidad', 1)
        unidad = getattr(item, 'unidad_medida', 'unidades')

        ws[f'A{fila_actual}'] = i
        ws[f'B{fila_actual}'] = nombre_completo
        ws[f'C{fila_actual}'] = cantidad
        ws[f'D{fila_actual}'] = unidad
        ws[f'E{fila_actual}'] = precio_estimado
        ws[f'F{fila_actual}'] = margen_base 
        
        # Fórmula de Excel (Cantidad * Costo * (1 + Margen))
        ws[f'G{fila_actual}'] = f"=C{fila_actual}*E{fila_actual}*(1+F{fila_actual})"
        
        # Formato moneda y porcentaje
        ws[f'E{fila_actual}'].number_format = '#,##0'
        ws[f'F{fila_actual}'].number_format = '0%'
        ws[f'G{fila_actual}'].number_format = '#,##0'

        # Aplicar bordes
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{fila_actual}'].border = borde_fino
        
        fila_actual += 1

    # --- TOTALES FINALES ---
    fila_actual += 1
    ws[f'F{fila_actual}'] = "Subtotal Neto:"
    # Aseguramos que la suma no falle si solo hay 1 ítem
    rango_suma = f"G8:G{fila_actual-2}" if fila_actual > 9 else "G8"
    ws[f'G{fila_actual}'] = f"=SUM({rango_suma})"
    ws[f'F{fila_actual}'].font = fuente_negrita
    
    fila_actual += 1
    ws[f'F{fila_actual}'] = "IVA (19%):"
    ws[f'G{fila_actual}'] = f"=G{fila_actual-1}*0.19"
    
    fila_actual += 1
    ws[f'F{fila_actual}'] = "TOTAL FINAL ($):"
    ws[f'G{fila_actual}'] = f"=G{fila_actual-2}+G{fila_actual-1}"
    ws[f'F{fila_actual}'].font = fuente_negrita
    ws[f'G{fila_actual}'].font = Font(bold=True, size=12)

    # --- SECCIÓN DE NOTAS / IA ---
    fila_actual += 2
    ws.merge_cells(f'A{fila_actual}:G{fila_actual}')
    ws[f'A{fila_actual}'] = "💡 Notas del Levantamiento Técnico"
    ws[f'A{fila_actual}'].font = fuente_negrita
    
    fila_actual += 1
    ws.merge_cells(f'A{fila_actual}:G{fila_actual + 1}')
    # Buscamos notas adicionales, si no hay, ponemos un mensaje estándar
    justificacion = getattr(datos, 'notas_adicionales', "Sin observaciones adicionales.")
    ws[f'A{fila_actual}'] = justificacion
    ws[f'A{fila_actual}'].alignment = Alignment(wrap_text=True, vertical="top")

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18

    # Guardar archivo
    wb.save(nombre_archivo)
    return os.path.abspath(nombre_archivo)

if __name__ == "__main__":
    print("Módulo de Excel optimizado y listo para ser usado por el sistema principal.")